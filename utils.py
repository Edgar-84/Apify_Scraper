import os
import json
import time
import requests

import openpyxl
from requests.models import Response

from config import params
from logger_settings import logger


def delete_files_in_temp():
    for filename in os.listdir(params.temp_dir):
        if filename == '.gitkeep':
            continue

        file_path = os.path.join(params.temp_dir, filename)
        try:
            if os.path.isfile(file_path):
                os.remove(file_path)
        except Exception as ex:
            logger.error(f"Mistake during delete file: {file_path}, mistake: {ex}")

    logger.info("Clearing of the Temp directory is finished!")


def dump_json(data: dict, json_name: str) -> bool:
    try:
        with open(json_name, "w", encoding='utf-8') as file:
            json.dump(data, file, indent=4, ensure_ascii=False)
            return True

    except Exception as ex:
        logger.error(f"Mistake during dump_json: {ex}")
        return False


def send_request(url: str,
                 headers: dict = None,
                 pause: float = params.pause_before_request,
                 count_repeat_request: int = 10) -> Response or None:
    """
    Get Response object for url, with pause before requests and repeat
    Return: None if catch mistake
    """

    mistake_pause = 0
    time.sleep(pause)

    try:
        response = requests.get(url=url, headers=headers)

        if 400 <= response.status_code <= 599:
            while mistake_pause != 10:
                time.sleep(mistake_pause)
                response = requests.get(url=url, headers=headers)
                if 400 <= response.status_code <= 599:
                    mistake_pause += 1
                    continue

                else:
                    return response

            error_message = (f"Catch mistake during load url: {url}, status_code: {response.status_code}"
                             f"mistake: {response.text}")
            logger.error(error_message)
            return None

        else:
            logger.info(f"The site {url} has been successfully loaded.")
            return response

    except Exception as ex:
        error_message = f"Catch mistake during load url: {url}, mistake: {ex}"
        logger.critical(error_message)
        return None


def save_html_page(url: str,
                   headers: dict,
                   path_save: str) -> bool:

    result = send_request(url=url, headers=headers)

    if result is None:
        return False

    try:
        with open(path_save, 'w', encoding="utf-8") as file:
            file.write(result.text)
            logger.info(f"URL {url} is done to html.page")
            return True

    except Exception as ex:
        logger.error(f"Mistake during save_html_page: {ex}")
        return False


def create_excel(data: dict, name_save: str):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    head_data = None
    all_data_values = []

    for i in data:
        head_data = data[i].keys()
        break

    for i in data:
        all_data_values.append(data[i].values())

    column = 1
    for key in head_data:
        sheet.cell(row=1, column=column).value = key
        column += 1

    column = 1
    row = 2
    for values in all_data_values:
        for value in values:
            sheet.cell(row=row, column=column).value = value
            column += 1
        row += 1
        column = 1

    workbook.save(name_save)


def prepared_dict(path_to_json: str):
    with open(path_to_json, encoding='utf-8') as file:
        data = json.load(file)
        return data

#
# data_dict = prepared_dict(path_to_json=params.temporary_json_all_press)
# create_excel(data_dict, 'First_24.xlsx')

